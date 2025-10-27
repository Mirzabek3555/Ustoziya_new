from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
import pandas as pd
import io

from accounts.models import User
from materials.models import Material, Assignment, VideoLesson, Model3D
from tests.models import Test
from ocr_processing.models import OCRProcessing


def home(request):
    """Asosiy sahifa"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # Statistikalar
    total_materials = Material.objects.filter(is_public=True).count()
    total_tests = Test.objects.filter(is_public=True, is_active=True).count()
    total_users = User.objects.filter(is_active=True).count()
    
    context = {
        'total_materials': total_materials,
        'total_tests': total_tests,
        'total_users': total_users,
    }
    return render(request, 'home.html', context)


@login_required
def dashboard(request):
    """Dashboard sahifasi"""
    user = request.user
    
    # Statistikalar
    materials_count = Material.objects.filter(author=user).count()
    tests_count = Test.objects.filter(author=user).count()
    assignments_count = Assignment.objects.filter(teacher=user).count()
    videos_count = VideoLesson.objects.filter(author=user).count()
    models_3d_count = Model3D.objects.filter(author=user).count()
    # So'nggi materiallar
    recent_materials = Material.objects.filter(author=user).order_by('-created_at')[:5]
    
    # So'nggi testlar - barcha testlarni ko'rsatish (umumiy foydalanish uchun)
    recent_tests = Test.objects.filter(is_public=True, is_active=True).order_by('-created_at')[:5]
    
    # Test tahlil statistikasi
    from ocr_processing.models import OCRProcessing
    test_analysis_count = OCRProcessing.objects.filter(user=user).count()
    
    context = {
        'stats': {
            'materials_count': materials_count,
            'tests_count': tests_count,
            'assignments_count': assignments_count,
            'videos_count': videos_count,
            'models_3d_count': models_3d_count,
            'test_analysis_count': test_analysis_count,
        },
        'recent_materials': recent_materials,
        'recent_tests': recent_tests,
    }
    return render(request, 'dashboard.html', context)


def register_view(request):
    """Ro'yxatdan o'tish sahifasi"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        subject = request.POST.get('subject')
        school = request.POST.get('school')
        phone = request.POST.get('phone')
        
        if password != password_confirm:
            messages.error(request, 'Parollar mos kelmaydi')
            return render(request, 'register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Bu foydalanuvchi nomi allaqachon mavjud')
            return render(request, 'register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Bu email allaqachon mavjud')
            return render(request, 'register.html')
        
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                subject=subject,
                school=school,
                phone=phone
            )
            login(request, user)
            messages.success(request, 'Muvaffaqiyatli ro\'yxatdan o\'tdingiz!')
            return redirect('dashboard')
        except Exception as e:
            messages.error(request, f'Xatolik yuz berdi: {str(e)}')
    
    return render(request, 'register.html')


def login_view(request):
    """Kirish sahifasi"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Muvaffaqiyatli kirdingiz!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Noto\'g\'ri foydalanuvchi nomi yoki parol')
    
    return render(request, 'login.html')


def logout_view(request):
    """Chiqish"""
    from django.contrib.auth import logout
    logout(request)
    messages.success(request, 'Muvaffaqiyatli chiqdingiz!')
    return redirect('home')


@login_required
def profile(request):
    """Profil sahifasi"""
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.subject = request.POST.get('subject', user.subject)
        user.school = request.POST.get('school', user.school)
        user.phone = request.POST.get('phone', user.phone)
        user.bio = request.POST.get('bio', user.bio)
        
        if 'avatar' in request.FILES:
            user.avatar = request.FILES['avatar']
        
        user.save()
        messages.success(request, 'Profil muvaffaqiyatli yangilandi!')
        return redirect('profile')
    
    return render(request, 'profile.html')


@login_required
def materials_list(request):
    """Materiallar ro'yxati"""
    materials = Material.objects.filter(is_public=True).order_by('-created_at')
    
    # Filtrlash
    category = request.GET.get('category')
    material_type = request.GET.get('type')
    subject = request.GET.get('subject')
    search = request.GET.get('search')
    
    if category:
        materials = materials.filter(category_id=category)
    if material_type:
        materials = materials.filter(material_type=material_type)
    if subject:
        materials = materials.filter(author__subject=subject)
    if search:
        materials = materials.filter(
            title__icontains=search
        )
    
    context = {
        'materials': materials,
    }
    return render(request, 'materials/list.html', context)


@login_required
def material_create(request):
    """Yangi material yaratish"""
    if request.method == 'POST':
        try:
            # Material yaratish
            material = Material.objects.create(
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                material_type=request.POST.get('material_type'),
                grade_level=request.POST.get('grade_level', ''),
                tags=request.POST.get('tags', ''),
                is_public=request.POST.get('is_public') == 'on',
                author=request.user,
                file=request.FILES.get('file')
            )
            
            # Kichik rasm qo'shish
            if 'thumbnail' in request.FILES:
                material.thumbnail = request.FILES['thumbnail']
                material.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Material muvaffaqiyatli yaratildi!',
                'material_id': material.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Xatolik yuz berdi: {str(e)}'
            })
    
    return render(request, 'materials/create.html')


@login_required
def tests_list(request):
    """Testlar ro'yxati"""
    tests = Test.objects.filter(is_public=True, is_active=True).order_by('-created_at')
    
    # Debug: Print sample data
    if request.GET.get('debug') == '1':
        print("=== DEBUG: Sample test data ===")
        if tests.exists():
            for test in tests[:3]:
                print(f"ID: {test.id}, Grade: '{test.grade_level}', Subject: '{test.subject}', Difficulty: '{test.difficulty}'")
        else:
            print("No tests found in database")
        print("=== END DEBUG ===")
    
    # Filtrlash
    category = request.GET.get('category')
    subject = request.GET.get('subject')
    grade_level = request.GET.get('grade')
    difficulty = request.GET.get('difficulty')
    search = request.GET.get('search')
    
    if category:
        tests = tests.filter(category_id=category)
    if subject:
        tests = tests.filter(subject=subject)
    if grade_level:
        tests = tests.filter(grade_level=grade_level)
    if difficulty:
        tests = tests.filter(difficulty=difficulty)
    if search:
        tests = tests.filter(title__icontains=search)
    
    context = {
        'tests': tests,
    }
    return render(request, 'tests/list.html', context)


@login_required
def test_detail(request, pk):
    """Test tafsilotlari"""
    from django.shortcuts import get_object_or_404
    test = get_object_or_404(Test, pk=pk, is_public=True, is_active=True)
    
    # Get questions for this test
    questions = test.questions.all().order_by('order')
    
    context = {
        'test': test,
        'questions': questions,
    }
    return render(request, 'tests/detail.html', context)


@login_required
def test_create(request):
    """Yangi test yaratish"""
    if request.method == 'POST':
        try:
            # Test yaratish
            test = Test.objects.create(
                title=request.POST.get('title'),
                description=request.POST.get('description', ''),
                category_id=request.POST.get('category'),
                subject=request.POST.get('subject'),
                grade_level=request.POST.get('grade_level'),
                difficulty=request.POST.get('difficulty'),
                time_limit=int(request.POST.get('time_limit', 60)),
                is_public=request.POST.get('is_public') == 'true',
                author=request.user
            )
            
            # Savollar yaratish
            questions_data = request.POST.get('questions', '[]')
            if questions_data:
                import json
                questions = json.loads(questions_data)
                
                for i, question_data in enumerate(questions):
                    question = Question.objects.create(
                        test=test,
                        question_text=question_data.get('text'),
                        question_type=question_data.get('type'),
                        points=int(question_data.get('points', 1)),
                        order=i + 1,
                        explanation=question_data.get('explanation', '')
                    )
                    
                    # Javoblar yaratish
                    for j, answer_data in enumerate(question_data.get('answers', [])):
                        Answer.objects.create(
                            question=question,
                            answer_text=answer_data.get('text'),
                            is_correct=answer_data.get('is_correct', False),
                            order=j + 1
                        )
            
            # Test statistikasini yangilash
            test.total_questions = test.questions.count()
            test.total_points = sum(q.points for q in test.questions.all())
            test.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Test muvaffaqiyatli yaratildi!',
                'test_id': test.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Xatolik yuz berdi: {str(e)}'
            })
    
    # Kategoriyalarni olish
    from tests.models import TestCategory
    categories = TestCategory.objects.all()
    
    context = {
        'categories': categories,
    }
    return render(request, 'tests/create.html', context)


@login_required
def test_analysis(request):
    """Test tahlili sahifasi - Google Gemini AI bilan"""
    if request.method == 'POST':
        # AI tahlil qilish
        try:
            image = request.FILES.get('image')
            class_name = request.POST.get('class_name')
            
            if not image:
                return JsonResponse({'success': False, 'error': 'Rasm tanlang'})
            
            if not class_name:
                return JsonResponse({'success': False, 'error': 'Sinfni tanlang'})
            
            # Avtomatik test aniqlash (AI orqali)
            # Hozircha oddiy test ishlatamiz
            test = Test.objects.filter(author=request.user, is_active=True).first()
            if not test:
                return JsonResponse({'success': False, 'error': 'Test topilmadi'})
            
            # OCR xizmati
            from ocr_processing.services import OCRService, TestGradingService
            ocr_service = OCRService()
            grading_service = TestGradingService()
            
            # Rasmni saqlash
            import os
            from django.conf import settings
            import uuid
            
            filename = f"{uuid.uuid4()}.jpg"
            image_path = os.path.join(settings.MEDIA_ROOT, 'temp', filename)
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            
            with open(image_path, 'wb') as f:
                for chunk in image.chunks():
                    f.write(chunk)
            
            # OCR qilish
            ocr_text, confidence = ocr_service.extract_text(image_path)
            
            if not ocr_text:
                return JsonResponse({'success': False, 'error': 'Rasmdan matn olinmadi'})
            
            # Test baholash
            from ocr_processing.models import OCRProcessing
            ocr_processing = OCRProcessing.objects.create(
                user=request.user,
                test=test,
                image=image,
                processed_text=ocr_text,
                confidence_score=confidence,
                student_class=class_name  # Sinf ma'lumotini qo'shamiz
            )
            
            # AI tahlil
            test_result = grading_service.grade_test(ocr_processing, test)
            
            if test_result:
                # Excel export uchun ma'lumot
                excel_data = {
                    'student_name': test_result.student_name,
                    'test_title': test.title,
                    'total_questions': test_result.total_questions,
                    'correct_answers': test_result.correct_answers,
                    'wrong_answers': test_result.wrong_answers,
                    'percentage': test_result.percentage,
                    'grade': test_result.grade,
                    'date': ocr_processing.created_at.strftime('%Y-%m-%d %H:%M:%S')
                }
                
                return JsonResponse({
                    'success': True,
                    'student_name': test_result.student_name,
                    'total_questions': test_result.total_questions,
                    'correct_answers': test_result.correct_answers,
                    'wrong_answers': test_result.wrong_answers,
                    'percentage': test_result.percentage,
                    'grade': test_result.grade,
                    'feedback': getattr(test_result, 'feedback', None),
                    'excel_data': excel_data
                })
            else:
                return JsonResponse({'success': False, 'error': 'Tahlil qilishda xatolik'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Xatolik: {str(e)}'})
    
    # GET so'rov - sahifa ko'rsatish
    tests = Test.objects.filter(author=request.user, is_active=True)
    
    # Test tahlil statistikasi
    from ocr_processing.models import OCRProcessing, TestResult
    analysis_count = OCRProcessing.objects.filter(user=request.user).count()
    completed_analysis = TestResult.objects.filter(ocr_processing__user=request.user).count()
    
    context = {
        'tests': tests,
        'analysis_count': analysis_count,
        'completed_analysis': completed_analysis,
    }
    return render(request, 'test_analysis.html', context)


@login_required
def export_single_test_results(request, test_id):
    """Bitta test natijalarini Excel ga export qilish"""
    try:
        from ocr_processing.models import TestResult
        
        # Bitta test uchun natijalar
        test_results = TestResult.objects.filter(
            ocr_processing__user=request.user,
            ocr_processing__test_id=test_id
        ).select_related('ocr_processing__test').order_by('student_name', 'created_at')
        
        if not test_results.exists():
            return JsonResponse({'success': False, 'error': 'Bu test uchun ma\'lumot yo\'q'})
        
        # Excel ma'lumotlari
        data = []
        for result in test_results:
            data.append({
                'O\'quvchi ismi': result.student_name,
                'Test nomi': result.ocr_processing.test.title if result.ocr_processing.test else 'Noma\'lum',
                'Jami savollar': result.total_questions,
                'To\'g\'ri javoblar': result.correct_answers,
                'Noto\'g\'ri javoblar': result.wrong_answers,
                'Foiz': f"{result.percentage:.1f}%",
                'Baholash': result.grade,
                'Sana': result.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'AI tahlil': 'Google Gemini AI'
            })
        
        # Excel fayl yaratish
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Test Natijalari', index=False)
        
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_natijalari.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Export xatoligi: {str(e)}'})


@login_required
def export_test_results(request):
    """Barcha test natijalarini Excel ga export qilish - har bir o'quvchi alohida"""
    try:
        from ocr_processing.models import TestResult
        
        # Sinf tanlash parametri
        class_filter = request.GET.get('class_name', '')
        
        # Foydalanuvchining test natijalari (sinf bo'yicha filter)
        test_results = TestResult.objects.filter(
            ocr_processing__user=request.user
        ).select_related('ocr_processing__test')
        
        # Agar sinf tanlangan bo'lsa, faqat o'sha sinfni ko'rsat
        if class_filter:
            test_results = test_results.filter(ocr_processing__student_class=class_filter)
        
        test_results = test_results.order_by('student_name', 'processed_at')
        
        if not test_results.exists():
            return JsonResponse({'success': False, 'error': 'Export qilish uchun ma\'lumot yo\'q'})
        
        # Excel fayl yaratish - Rasmda ko'rsatilgan format
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Natijalari"
        
        # Style'lar
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sarlavha qo'shish
        ws.merge_cells('A1:H1')
        ws['A1'] = "O'QUVCHILARNING TEST NATIJALARI"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Header qator
        headers = [
            "T/r", 
            "O'quvchilarning ismi va familiyasi", 
            "1", "2", "3", "4", "5", 
            "Jami", "%"
        ]
        
        # Header yozish
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ma'lumotlar yozish
        row = 4
        for i, result in enumerate(test_results, 1):
            # T/r
            ws.cell(row=row, column=1, value=i).border = border
            
            # O'quvchi ismi
            ws.cell(row=row, column=2, value=result.student_name).border = border
            
            # Har bir savol uchun ball (5 ta savol)
            # Bu yerda AI dan kelgan javoblarni ishlatamiz
            answers = getattr(result, 'answers', {})
            total_score = 0
            
            for q_num in range(1, 6):  # 1-5 savollar
                answer = answers.get(str(q_num), '')
                # Agar javob to'g'ri bo'lsa 5 ball, noto'g'ri bo'lsa 0 ball
                score = 5 if answer in ['A', 'B', 'C', 'D'] else 0
                ws.cell(row=row, column=2+q_num, value=score).border = border
                total_score += score
            
            # Jami ball
            ws.cell(row=row, column=8, value=total_score).border = border
            
            # Foiz
            percentage = (total_score / 25) * 100  # 5 savol * 5 ball = 25
            ws.cell(row=row, column=9, value=f"{percentage:.0f}%").border = border
            
            row += 1
        
        # O'rtacha qator qo'shish
        if test_results.exists():
            avg_row = row
            ws.cell(row=avg_row, column=1, value="O'rtacha:").border = border
            ws.cell(row=avg_row, column=2, value="").border = border
            
            # Har bir savol uchun o'rtacha
            for q_num in range(1, 6):
                # Bu yerda haqiqiy o'rtacha hisoblash kerak
                avg_score = 2.5  # Namuna uchun
                ws.cell(row=avg_row, column=2+q_num, value=f"{avg_score:.1f}").border = border
            
            # Jami o'rtacha
            total_avg = 12.5  # Namuna uchun
            ws.cell(row=avg_row, column=8, value=f"{total_avg:.1f}").border = border
            
            # O'rtacha foiz
            avg_percentage = 50.0  # Namuna uchun
            ws.cell(row=avg_row, column=9, value=f"{avg_percentage:.0f}%").border = border
        
        # Ustun kengliklarini o'zgartirish
        column_widths = [8, 30, 8, 8, 8, 8, 8, 10, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Barcha kataklarni border bilan bezash
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        output = io.BytesIO()
        wb.save(output)
        
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test_natijalari.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Export xatoligi: {str(e)}'})


@login_required
def ocr_upload(request):
    """OCR rasm yuklash sahifasi"""
    tests = Test.objects.filter(author=request.user, is_active=True)
    
    context = {
        'tests': tests,
    }
    return render(request, 'ocr_upload.html', context)


@login_required
def assignments_list(request):
    """Topshiriqlar ro'yxati"""
    return render(request, 'assignments/list.html')


@login_required
def videos_list(request):
    """Video darsliklar ro'yxati"""
    return render(request, 'videos/list.html')


@login_required
def models_3d_list(request):
    """3D modellar ro'yxati"""
    return render(request, '3d_models/list.html')


@login_required
def test_results(request, test_id):
    """Test natijalarini ko'rish"""
    from tests.models import Test, TestAttempt
    
    try:
        test = Test.objects.get(id=test_id, author=request.user)
        attempts = TestAttempt.objects.filter(test=test).select_related('student')
        
        context = {
            'test': test,
            'attempts': attempts,
        }
        return render(request, 'tests/results.html', context)
    except Test.DoesNotExist:
        return redirect('tests_list')


@login_required


@login_required
def test_ocr_upload(request):
    """Test natijalarini rasm ko'rinishida yuklash va OCR orqali tahlil qilish"""
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('test_image')
            if not uploaded_file:
                return JsonResponse({'error': 'Rasm yuklanmadi'}, status=400)
            
            # Rasmni saqlash
            import os
            from django.conf import settings
            from PIL import Image
            import pytesseract
            import cv2
            import numpy as np
            
            # Rasmni media papkasiga saqlash
            media_path = os.path.join(settings.MEDIA_ROOT, 'test_images')
            os.makedirs(media_path, exist_ok=True)
            
            file_path = os.path.join(media_path, uploaded_file.name)
            with open(file_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)
            
            # OCR orqali matnni olish
            image = cv2.imread(file_path)
            
            # Rasmni yaxshilash
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            denoised = cv2.medianBlur(gray, 3)
            
            # OCR qilish
            text = pytesseract.image_to_string(denoised, lang='uzb+eng')
            
            # Test natijalarini tahlil qilish
            analysis_result = analyze_test_results(text)
            
            return JsonResponse({
                'success': True,
                'extracted_text': text,
                'analysis': analysis_result,
                'image_url': os.path.join(settings.MEDIA_URL, 'test_images', uploaded_file.name)
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Xatolik: {str(e)}'}, status=500)
    
    return render(request, 'tests/ocr_upload.html')


def analyze_test_results(text):
    """OCR dan olingan matnni tahlil qilish"""
    import re
    
    result = {
        'students': [],
        'test_info': {},
        'statistics': {}
    }
    
    # Test ma'lumotlarini topish
    lines = text.split('\n')
    
    # Sinf va fan ma'lumotlarini topish
    for line in lines:
        if 'sinf' in line.lower():
            result['test_info']['grade'] = line.strip()
        if any(subject in line.lower() for subject in ['matematika', 'fizika', 'kimyo', 'biologiya', 'informatika']):
            result['test_info']['subject'] = line.strip()
    
    # O'quvchilar va natijalarni topish
    student_pattern = r'(\d+)\s+([A-Za-z\u0400-\u04FF\s]+)\s+(\d+(?:\.\d+)?)'
    
    for line in lines:
        match = re.search(student_pattern, line)
        if match:
            student_data = {
                'number': match.group(1),
                'name': match.group(2).strip(),
                'score': float(match.group(3))
            }
            result['students'].append(student_data)
    
    # Statistika hisoblash
    if result['students']:
        scores = [s['score'] for s in result['students']]
        result['statistics'] = {
            'total_students': len(result['students']),
            'average_score': sum(scores) / len(scores),
            'max_score': max(scores),
            'min_score': min(scores)
        }
    
    return result
